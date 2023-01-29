import os
import sys
from PyQt5.QtCore import QEventLoop
from PyQt5.QtWidgets import *
from PyQt5.QAxContainer import *
from PyQt5.QtCore import QEventLoop, QSize
import time
import pandas as pd
import requests
import exchange_calendars as ecals
import datetime
from openpyxl import load_workbook
import random
import numpy as np
import urllib.request
from bs4 import BeautifulSoup

class MyWindow(QMainWindow):

    delay = 3.6
    def __init__(self):

        super().__init__()

        # 종료 메세지 출력
        self.setMinimumSize(QSize(500, 100))
        self.setWindowTitle("Alert Message")

        pybutton = QPushButton('버튼 눌러야 컴퓨터 계속 사용 가능', self)
        pybutton.clicked.connect(self.clickMethod)
        pybutton.resize(400, 30)
        pybutton.move(50, 35)

        self.kiwoom = QAxWidget("KHOPENAPI.KHOpenAPICtrl.1")
### event ###
        self.kiwoom.OnEventConnect.connect(self.event_connect)
        self.kiwoom.OnReceiveTrCondition.connect(self.receive_trCondition)
        self.kiwoom.OnReceiveConditionVer.connect(self.receive_VerCondition)

    def clickMethod(self):
        QMessageBox.about(self, "", "자동종료 취소")
        sys.exit()

    def receive_VerCondition(self, ret, msg):
        self.receive_VerCondition.exit()

    def receive_trCondition(self, screenNo, codes, conditionName, conditionIndex, inquiry):
        print("@@ trCondition 시작")
        if codes == "":
            print("@@ 검색결과 없음")
        self.codeList = codes.split(';')
        del self.codeList[-1]
        self.receive_trCondition_loop.exit()

    def event_connect(self, err_code):
        if err_code == 0:
            print("@@ 로그인 성공")
        self.login_event_loop.exit()

### 한국거래소 영업일 계산
    def calculateBusinessDay(self, day):
        # day = x 일때, x 일전 영업일을 출력
        day = day * -1
        XKRX = ecals.get_calendar("XKRX")  # 한국 코드

        count = 0
        targetDay = 0
        dayCount = day + 2

        for x in range(dayCount - 1):
            while targetDay != x + 1:
                if XKRX.is_session(datetime.date.today() - datetime.timedelta(days=count)):
                    targetDay = targetDay + 1
                count = count + 1
            todayDay = str(datetime.date.today() - datetime.timedelta(days=count - 1)).replace("-", "")

        return todayDay

### 실행 함수###
    def kiwoom_login(self):
        print("@@ 로그인 시작")
        self.login_event_loop = QEventLoop()
        self.kiwoom.CommConnect()
        self.login_event_loop.exec_()
    
    # 조건식 불러오기
    def load_conditionList(self):
        print("@@ 조건식 가져오기 시작")
        self.receive_VerCondition = QEventLoop()

        isLoad = self.kiwoom.GetConditionLoad()
        self.receive_VerCondition.exec_()
        conditions = self.kiwoom.dynamicCall("GetConditionNameList()")
        if isLoad == 1:
            print("@@ 조건식 가져오기 성공")
        else:
            print("@@ 조건식 가져오기 실패")
        time.sleep(self.delay)
        return conditions
    
    # 조건식 검색, 코드리스트 불러오기
    def search_condition(self, conditionName, conditionIdx):
        print("@@ 조건검색 시작")

        self.receive_trCondition_loop = QEventLoop()
        screenNo = "0150"
        isRealTime = 0
        isRequest = self.kiwoom.dynamicCall("SendCondition(QString, QString, int, int)",
                                            screenNo, conditionName, conditionIdx, isRealTime)
        self.receive_trCondition_loop.exec_()
        if isRequest == 0:
            print("@@ 조건검색 실패")
        if isRequest == 1:
            print("@@ 조건검색 성공")
        time.sleep(self.delay)
        return self.codeList

    # 주식정보 불러오기
    def getStockInfo(self, code, pages):

        df = pd.DataFrame()
        for page in range(1, pages):
            url = "https://finance.naver.com/item/sise_day.naver?code={}&page={}".format(code, page)
            res = requests.get(url, headers={'User-agent': 'Mozilla/5.0'})
            df = pd.concat([df, pd.read_html(res.text, header=0)[0]], axis=0)
            # df['저가-시가'] = df['저가'] - df['시가']
            # df['고가-시가'] = df['고가'] - df['시가']

        # df.dropna()를 이용해 결측값 있는 행 제거
        df = df.dropna()
        return df

    # 전체 종목코드 가져오기
    def GetCodeListByMarket(self, market):
        codeList = self.kiwoom.dynamicCall("GetCodeListByMarket(QString)", market)
        return codeList

    # 코스피, 코스닥정보 불러오기
    def getKOSPIInfo(self):

        basic_url = "https://finance.naver.com/sise/sise_index_day.naver?code=KOSPI&page=1"
        fp = urllib.request.urlopen(basic_url)
        source = fp.read()
        fp.close()

        soup = BeautifulSoup(source, 'html.parser')
        soup = soup.findAll("span", class_="tah p11 red01")

        value = soup[0].string.replace(" ", "").strip()

        return value

    def getKOSDAQInfo(self):

        basic_url = "https://finance.naver.com/sise/sise_index_day.naver?code=KOSDAQ&page=1"
        fp = urllib.request.urlopen(basic_url)
        source = fp.read()
        fp.close()

        soup = BeautifulSoup(source, 'html.parser')
        soup = soup.findAll("span", class_="tah p11 red01")

        value = soup[0].string.replace(" ", "").strip()

        return value

    # 종목코드로 종목명 불러오기
    def GetMasterCodeName(self, code):
        code_name = self.kiwoom.dynamicCall("GetMasterCodeName(QString)", code)
        return code_name

    # 종목 매수 (in_strAccount, in_strCode, out_result)

    # 종목 매도 (in_strAccount, in_strCode, out_result)

    # 보유 종목 출력 (in_strAccount, out_listCodes)


    # 주식정보 엑셀 저장하기
    def saveDataframeToExcel(self, df, fileName, sheetName):
        if os.path.isfile(fileName):
            ExcelWorkbook = load_workbook(fileName)
            if sheetName in ExcelWorkbook.sheetnames:
                ExcelWorkbook.remove(ExcelWorkbook[sheetName])
            # Generating the writer engine
            writer = pd.ExcelWriter(fileName, engine='openpyxl')
            # Assigning the workbook to the writer engine
            writer.book = ExcelWorkbook
            df.to_excel(writer, index=False, sheet_name=sheetName)
            writer.save()
            writer.close()
        else:
            df.to_excel(fileName, index=False, sheet_name=sheetName)
        # 주식정보 비교하기

    # 엑셀불러오기
    def readExcelToDataFrame(self, fileName, sheetName):
        global df
        if os.path.isfile(fileName):
            df = pd.read_excel(fileName, dtype=str, sheet_name=sheetName)
        else:
            pass
        return df

    def compare_stockInfo(self, code):
        pass

    # 특정 시트에 종목 넣기
    def insert_StockToSheet(self, code, sheet):
        pass

    # 날짜별 수익률 계산 (인수 : DF)
    def ReturnMoneyPerDay(self, df):
        dfDate = df.copy().drop_duplicates(subset=['날짜'], keep='last')
        lstDate = []
        dfResult = pd.DataFrame([], columns=['날짜3', '총수익률', '수익종목수', '손해종목수'])
        for index, row in dfDate.iterrows():
            lstDate.append(row['날짜'])
        for i in lstDate:
            # 해당 날짜 분류
            dfSort = df.loc[df['날짜'] == i].copy()
            # 결측값 제거
            dfSort['수익률'].replace('', np.nan, inplace=True)
            dfSort = dfSort.dropna(subset=['수익률'])
            dfSort['수익률'] = dfSort['수익률'].astype(int)
            # dfSort = dfSort['수익률']
            Total = dfSort['수익률'].sum()
            benefit = dfSort.loc[dfSort['수익률'] == 1, '수익률'].count()
            loss = dfSort.loc[dfSort['수익률'] == -5, '수익률'].count()

            newRow = {'날짜3': [i],
                      '총수익률': [Total],
                      '수익종목수': [benefit],
                      '손해종목수': [loss]}
            newResult = pd.DataFrame(data=newRow)
            dfResult = pd.concat([dfResult, newResult], ignore_index=True)
        KOSPIUpDown = myWindow.getKOSPIInfo()
        KOSDAQUpDown = myWindow.getKOSDAQInfo()
        newRow = {'날짜3': ['KOSPI'],
                  '총수익률': [KOSPIUpDown],
                  '수익종목수': ['KOSDAQ'],
                  '손해종목수': [KOSDAQUpDown]}
        newResult = pd.DataFrame(data=newRow)
        dfResult = pd.concat([dfResult, newResult], ignore_index=True)
        return dfResult
    
    # 종료 메세지 박스 띄우기
    def quitDialog(self):
        pass

    # 코드리스트 필터링
    def filtered_code(self,list_code):
        new_code_list = []
        for code in list_code:
            codeName = self.GetMasterCodeName(code)
            # 스팩주 제거
            if "스팩" in codeName or "4호" in codeName:
                pass
            elif "ETN" in codeName:
                pass
            elif codeName[-1:] == "우" or "3우C" in codeName or "우B" in codeName or "G3우" in codeName:
                pass
            elif "TIGER" in codeName or "KOSEF" in codeName or "KBSTAR" in codeName or "KODEX" in codeName or "KINDEX" in codeName or "TREX" in codeName:
                pass
            elif "ARIRANG" in codeName or "SMART" in codeName or "FOCUS" in codeName or "HANARO" in codeName or "TIMEFOLIO" in codeName or "네비게이터" in codeName:
                pass
            else:
                new_code_list.append(code)
        return new_code_list

    # 조건식리스트 필터링
    def filtered_condition(self, conditionList):

        new_condition_list = []
        for condition in conditionList:
            conditionNumber = condition.split("^")[0]
            conditionName = condition.split("^")[1]
            # 스팩주 제거
            if "종가1퍼상승마감" in conditionName or "종가3퍼상승마감" in conditionName or "종가5퍼상승마감" in conditionName or "종가10퍼상승마감" in conditionName \
                    or "종가1퍼하락마감" in conditionName or "종가3퍼하락마감" in conditionName or "종가5퍼하락마감" in conditionName or "종가10퍼하락마감" in conditionName \
                    or "시가1퍼상승" in conditionName or "시가3퍼상승" in conditionName or "시가5퍼상승" in conditionName or "시가10퍼상승" in conditionName\
                    or "시가1퍼하락" in conditionName or "시가3퍼하락" in conditionName or "시가5퍼하락" in conditionName or "시가10퍼하락" in conditionName\
                    or "고가1퍼상승" in conditionName or "고가3퍼상승" in conditionName or "고가5퍼상승" in conditionName or "고가10퍼상승" in conditionName\
                    or "고가1퍼하락" in conditionName or "고가3퍼하락" in conditionName or "고가5퍼하락" in conditionName or "고가10퍼하락" in conditionName\
                    or "저가1퍼상승" in conditionName or "저가3퍼상승" in conditionName or "저가5퍼상승" in conditionName or "저가10퍼상승" in conditionName\
                    or "저가1퍼하락" in conditionName or "저가3퍼하락" in conditionName or "저가5퍼하락" in conditionName or "저가10퍼하락" in conditionName:
                pass
            else:
                new_condition_list.append('^'.join([conditionNumber, conditionName]))

        return new_condition_list

if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        myWindow = MyWindow()
        folderPath = "C:/Users/zzang/Desktop/주식자료"
        TodayBusDay = str(myWindow.calculateBusinessDay(0))
        YesterdayBusDay = str(myWindow.calculateBusinessDay(-1))
        ResultFileName = "C:/Users/zzang/Desktop/주식자료/Result.xlsx"
        conditionName = ""
        processStage = ""

        myWindow.show()
        # 키움 로그인
        myWindow.kiwoom_login()

        # 조건식 불러오기
        conditionList = myWindow.load_conditionList().split(";")
        del conditionList[-1]
        conditionListCopy = conditionList
        conditionList = []

        # 조건식 이름 변경
        processStage = "조건식 이름변경"
        for name in conditionListCopy:
            name = name.replace("/","")
            conditionList.append(name)

        # 조건식 필터링
        conditionList = myWindow.filtered_condition(conditionList)

        for i in conditionList:
            print(i)

        # 조건식별 반복작업
        processStage = "조건식 반복작업"
        for condition in conditionList:
            #if condition.split("^")[1] != "N자눌림목":
            #if condition.split("^")[1] != "5일전매수":
                #continue
            fileName = os.path.join(folderPath, str(condition.split("^")[1])) + ".xlsx"
            sheetName = TodayBusDay
            conditionName = condition.split("^")[1]

        # 금일 주식정보 저장
            processStage = "금일 주식정보 저장"
            # 조건식 종목 리스트 불러오기
            list_code = myWindow.search_condition(conditionName, condition.split("^")[0])

            # 코드 리스트 필터링
            list_code = myWindow.filtered_code(list_code)

            new_df = pd.DataFrame(columns=['code', '시가', '종가', '고가', '저가', '날짜', '시가2', '종가2', '고가2', '저가2', '날짜2', '결과', '매도날짜', '수익률'])
            print("@@ 금일 주식정보 저장 : {0}, {1}".format(conditionName, list_code))
            len_list_code = len(list_code)
            if len(list_code) > 100:
                list_code = random.sample(list_code, 100)
            for idx, code in enumerate(list_code):
                # 주식정보 불러오기 (네이버 크롤링)
                print("code : {0} -- {1} / {2}".format(code, len_list_code, idx+1))
                df = myWindow.getStockInfo(code, 2)
                # 새 데이터 프레임에 주식정보 저장
                new_df = new_df.append(
                    {'code': code,
                     '시가': df.at[1, '시가'],
                     '종가': df.at[1, '종가'],
                     '고가': df.at[1, '고가'],
                     '저가': df.at[1, '저가'],
                     '날짜': df.at[1, '날짜'],
                     '시가2': '',
                     '종가2': '',
                     '고가2': '',
                     '저가2': '',
                     '날짜2': '',
                     '결과': '',
                     '매도날짜': '',
                     '수익률': ''}, ignore_index=True)

            #조건식별 데이터 프레임 엑셀 저장
            myWindow.saveDataframeToExcel(new_df, fileName, sheetName)

            if os.path.isfile(fileName):
                ExcelWorkbook = load_workbook(fileName)

                # Result 시트 작성, 매입 된 주식들 매도 판단
                sheetName = "Result"
                if sheetName in ExcelWorkbook.sheetnames:
                    dfResult = pd.read_excel(fileName, dtype=str, sheet_name=sheetName)
                # df 코드번호 반복작업

                    for index, row in dfResult.iterrows():
                        if not str(dfResult.loc[index, '수익률']) == 'nan':
                            continue
                        df_stock = myWindow.getStockInfo(row['code'], 2)
                        dfResult.loc[index, '종가2'] = df_stock.at[1, '종가']
                        dfResult.loc[index, '시가2'] = df_stock.at[1, '시가']
                        dfResult.loc[index, '고가2'] = df_stock.at[1, '고가']
                        dfResult.loc[index, '저가2'] = df_stock.at[1, '저가']
                        dfResult.loc[index, '날짜2'] = df_stock.at[1, '날짜']
                    # 매입여부 결정
                        if round(int(row['종가']) * 0.985 * 0.95, 0) > int(df_stock.at[1, '저가']):
                            dfResult.loc[index, '수익률'] = "-5"
                        if round(int(row['종가']) * 0.985 * 1.015, 0) < int(df_stock.at[1, '고가']):
                            dfResult.loc[index, '수익률'] = "+1"
                        else:
                            pass
                    myWindow.saveDataframeToExcel(dfResult, fileName, sheetName)
                else:
                    dfResult = pd.DataFrame([], columns=['code', '종가', '시가', '고가', '저가', '날짜', '종가2', '시가2', '고가2', '저가2',
                                                         '날짜2', '결과', '매도날짜', '수익률'])
                    myWindow.saveDataframeToExcel(dfResult, fileName, sheetName)

                # 전 영업일 시트 데이터 금일자 데이터로 비교

                sheetName = YesterdayBusDay

                if sheetName in ExcelWorkbook.sheetnames:
                    df = pd.read_excel(fileName, dtype=str, sheet_name=sheetName)

                    # df 코드번호 반복작업
                    for index, row in df.iterrows():
                        # print(index, df.loc[index, 'code'])
                        df_stock = myWindow.getStockInfo(row['code'], 2)

                        # 1행 미존재 경우 다음 코드 진행
                        if str(df_stock.iat[0, 0]).replace(".","") != TodayBusDay:
                            continue

                        df.loc[index, '종가2'] = df_stock.at[1, '종가']
                        df.loc[index, '시가2'] = df_stock.at[1, '시가']
                        df.loc[index, '고가2'] = df_stock.at[1, '고가']
                        df.loc[index, '저가2'] = df_stock.at[1, '저가']
                        df.loc[index, '날짜2'] = df_stock.at[1, '날짜']

                        # 매입여부 결정, 해당 코드가 매입조건에 해당하면 dfResult 에 append 진행
                        if round(int(row['종가']) * 0.985, 0) > int(df_stock.at[1, '저가']):
                            # print(99)
                            df.loc[index, '결과'] = '매입'
                            # 매입한 코드만 Result 시트에 저장
                            newRow = {'code': [row['code']],
                                      '종가': [row['종가']],
                                      '시가': [row['시가']],
                                      '고가': [row['고가']],
                                      '저가': [row['저가']],
                                      '날짜': [row['날짜']],
                                      '종가2': [df_stock.at[1, '종가']],
                                      '시가2': [df_stock.at[1, '시가']],
                                      '고가2': [df_stock.at[1, '고가']],
                                      '저가2': [df_stock.at[1, '저가']],
                                      '날짜2': [df_stock.at[1, '날짜']],
                                      '결과': ['매입'],
                                      '매도날짜': [''],
                                      '수익률': [''],}
                            newResult = pd.DataFrame(data=newRow)
                            dfResult = pd.concat([dfResult, newResult], ignore_index=True)
                    dfResult = dfResult.drop_duplicates(subset=['code', '날짜', '날짜2'], keep='last')
                    # 일별 수익률 입력
                    dfTotal = pd.DataFrame([], columns=['날짜3', '총수익률', '수익종목수', '손해종목수'])
                    dfTotal = myWindow.ReturnMoneyPerDay(dfResult)

                    # 데이터프레임 엑셀저장
                    myWindow.saveDataframeToExcel(df, fileName, sheetName) # 전일자 저장
                    myWindow.saveDataframeToExcel(dfResult, fileName, "Result") # 결과시트 저장
                    myWindow.saveDataframeToExcel(dfTotal, ResultFileName, conditionName) # 결과파일 저장

    except Exception as e:

        print("{0} 에러, 조건식 : {1}".format(processStage, conditionName))
        print(e)

    # app.exec_()

