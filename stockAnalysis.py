import stockAnalysis
import os
import sys
from PyQt5.QtCore import QEventLoop
from PyQt5.QtWidgets import *
from PyQt5.QAxContainer import *
from PyQt5.QtCore import QEventLoop, QSize
import time
import pandas as pd
import requests
import exchange_calendars as ecals
import datetime
from openpyxl import load_workbook
import random
import numpy as np
import urllib.request
from bs4 import BeautifulSoup
import matplotlib.pyplot as plot


class MyWindow(QMainWindow):
    delay = 1

    def __init__(self):

        super().__init__()

        # 종료 메세지 출력
        self.setMinimumSize(QSize(500, 100))
        self.setWindowTitle("Alert Message")

        pybutton = QPushButton('버튼 눌러야 컴퓨터 계속 사용 가능', self)
        pybutton.clicked.connect(self.clickMethod)
        pybutton.resize(400, 30)
        pybutton.move(50, 35)

        self.kiwoom = QAxWidget("KHOPENAPI.KHOpenAPICtrl.1")
        ### event ###
        self.kiwoom.OnEventConnect.connect(self.event_connect)
        self.kiwoom.OnReceiveTrCondition.connect(self.receive_trCondition)
        self.kiwoom.OnReceiveConditionVer.connect(self.receive_VerCondition)

    def clickMethod(self):
        QMessageBox.about(self, "", "자동종료 취소")
        sys.exit()

    def receive_VerCondition(self, ret, msg):
        self.receive_VerCondition.exit()

    def receive_trCondition(self, screenNo, codes, conditionName, conditionIndex, inquiry):
        print("@@ trCondition 시작")
        if codes == "":
            print("@@ 검색결과 없음")
        self.codeList = codes.split(';')
        del self.codeList[-1]
        self.receive_trCondition_loop.exit()

    def event_connect(self, err_code):
        if err_code == 0:
            print("@@ 로그인 성공")
        self.login_event_loop.exit()

    ### 한국거래소 영업일 계산
    def calculateBusinessDay(self, day):
        # day = x 일때, x 일전 영업일을 출력
        day = day * -1
        XKRX = ecals.get_calendar("XKRX")  # 한국 코드

        count = 0
        targetDay = 0
        dayCount = day + 2

        for x in range(dayCount - 1):
            while targetDay != x + 1:
                if XKRX.is_session(datetime.date.today() - datetime.timedelta(days=count)):
                    targetDay = targetDay + 1
                count = count + 1
            todayDay = str(datetime.date.today() - datetime.timedelta(days=count - 1)).replace("-", "")

        return todayDay

    ### 실행 함수###
    def kiwoom_login(self):
        print("@@ 로그인 시작")
        self.login_event_loop = QEventLoop()
        self.kiwoom.CommConnect()
        self.login_event_loop.exec_()

    # 조건식 불러오기
    def load_conditionList(self):
        print("@@ 조건식 가져오기 시작")
        self.receive_VerCondition = QEventLoop()

        isLoad = self.kiwoom.GetConditionLoad()
        self.receive_VerCondition.exec_()
        conditions = self.kiwoom.dynamicCall("GetConditionNameList()")
        if isLoad == 1:
            print("@@ 조건식 가져오기 성공")
        else:
            print("@@ 조건식 가져오기 실패")
        time.sleep(self.delay)
        return conditions

    # 조건식 검색, 코드리스트 불러오기
    def search_condition(self, conditionName, conditionIdx):
        print("@@ 조건검색 시작")

        self.receive_trCondition_loop = QEventLoop()
        screenNo = "0150"
        isRealTime = 0
        isRequest = self.kiwoom.dynamicCall("SendCondition(QString, QString, int, int)",
                                            screenNo, conditionName, conditionIdx, isRealTime)
        self.receive_trCondition_loop.exec_()
        if isRequest == 0:
            print("@@ 조건검색 실패")
        if isRequest == 1:
            print("@@ 조건검색 성공")
        time.sleep(self.delay)
        return self.codeList

    # 주식정보 불러오기
    def getStockInfo(self, code, pages):

        df = pd.DataFrame()
        for page in range(1, pages):
            url = "https://finance.naver.com/item/sise_day.naver?code={}&page={}".format(code, page)
            res = requests.get(url, headers={'User-agent': 'Mozilla/5.0'})
            df = pd.concat([df, pd.read_html(res.text, header=0)[0]], axis=0)
            # df['저가-시가'] = df['저가'] - df['시가']
            # df['고가-시가'] = df['고가'] - df['시가']

        # df.dropna()를 이용해 결측값 있는 행 제거
        df = df.dropna()
        return df

    # 전체 종목코드 가져오기
    def GetCodeListByMarket(self, market):
        codeList = self.kiwoom.dynamicCall("GetCodeListByMarket(QString)", market)
        return codeList

    # 종목코드로 종목명 불러오기
    def GetMasterCodeName(self, code):
        code_name = self.kiwoom.dynamicCall("GetMasterCodeName(QString)", code)
        return code_name

    # 종목 매수 (in_strAccount, in_strCode, out_result)

    # 종목 매도 (in_strAccount, in_strCode, out_result)

    # 보유 종목 출력 (in_strAccount, out_listCodes)

    # 주식정보 엑셀 저장하기
    def saveDataframeToExcel(self, df, fileName, sheetName):
        if os.path.isfile(fileName):
            ExcelWorkbook = load_workbook(fileName)
            if sheetName in ExcelWorkbook.sheetnames:
                ExcelWorkbook.remove(ExcelWorkbook[sheetName])
            # Generating the writer engine
            writer = pd.ExcelWriter(fileName, engine='openpyxl')
            # Assigning the workbook to the writer engine
            writer.book = ExcelWorkbook
            df.to_excel(writer, index=False, sheet_name=sheetName)
            writer.save()
            writer.close()
        else:
            df.to_excel(fileName, index=False, sheet_name=sheetName)
        # 주식정보 비교하기

    # 엑셀불러오기
    def readExcelToDataFrame(self, fileName, sheetName):
        global df
        if os.path.isfile(fileName):
            df = pd.read_excel(fileName, dtype=str, sheet_name=sheetName)
        else:
            pass
        return df

    def compare_stockInfo(self, code):
        pass

    # 특정 시트에 종목 넣기
    def insert_StockToSheet(self, code, sheet):
        pass

    # 날짜별 수익률 계산 (인수 : DF)
    def ReturnMoneyPerDay(self, df):
        dfDate = df.copy().drop_duplicates(subset=['날짜'], keep='last')
        lstDate = []
        dfResult = pd.DataFrame([], columns=['날짜3', '총수익률', '수익종목수', '손해종목수'])
        for index, row in dfDate.iterrows():
            lstDate.append(row['날짜'])
        for i in lstDate:
            # 해당 날짜 분류
            dfSort = df.loc[df['날짜'] == i].copy()
            # 결측값 제거
            dfSort['수익률'].replace('', np.nan, inplace=True)
            dfSort = dfSort.dropna(subset=['수익률'])
            dfSort['수익률'] = dfSort['수익률'].astype(int)
            # dfSort = dfSort['수익률']
            Total = dfSort['수익률'].sum()
            benefit = dfSort.loc[dfSort['수익률'] == 1, '수익률'].count()
            loss = dfSort.loc[dfSort['수익률'] == -5, '수익률'].count()

            newRow = {'날짜3': [i],
                      '총수익률': [Total],
                      '수익종목수': [benefit],
                      '손해종목수': [loss]}
            newResult = pd.DataFrame(data=newRow)
            dfResult = pd.concat([dfResult, newResult], ignore_index=True)
        KOSPIUpDown = myWindow.getKOSPIInfo()
        KOSDAQUpDown = myWindow.getKOSDAQInfo()
        newRow = {'날짜3': ['KOSPI'],
                  '총수익률': [KOSPIUpDown],
                  '수익종목수': ['KOSDAQ'],
                  '손해종목수': [KOSDAQUpDown]}
        newResult = pd.DataFrame(data=newRow)
        dfResult = pd.concat([dfResult, newResult], ignore_index=True)
        return dfResult

    # 종료 메세지 박스 띄우기
    def quitDialog(self):
        pass

    # 코드리스트 필터링
    def filtered_code(self, list_code):
        new_code_list = []
        for code in list_code:
            codeName = self.GetMasterCodeName(code)
            # 스팩주 제거
            if "스팩" in codeName or "4호" in codeName:
                pass
            elif "ETN" in codeName:
                pass
            elif codeName[-1:] == "우" or "3우C" in codeName or "우B" in codeName or "G3우" in codeName:
                pass
            elif "TIGER" in codeName or "KOSEF" in codeName or "KBSTAR" in codeName or "KODEX" in codeName or "KINDEX" in codeName or "TREX" in codeName:
                pass
            elif "ARIRANG" in codeName or "SMART" in codeName or "FOCUS" in codeName or "HANARO" in codeName or "TIMEFOLIO" in codeName or "네비게이터" in codeName:
                pass
            else:
                new_code_list.append(code)

        return new_code_list


if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        myWindow = MyWindow()
        conditionName = ""
        processStage = ""

        myWindow.show()
        # 키움 로그인
        myWindow.kiwoom_login()

        # 조건식 불러오기
        conditionList = myWindow.load_conditionList().split(";")
        del conditionList[-1]
        conditionListCopy = conditionList
        conditionList = []

        # 조건식 이름 변경
        processStage = "조건식 이름변경"
        for name in conditionListCopy:
            name = name.replace("/", "")
            conditionList.append(name)

        # 총종목수 가져오기
        kospi = myWindow.GetCodeListByMarket('0').split(';')
        kosdaq = myWindow.GetCodeListByMarket('10').split(';')

        kospi = myWindow.filtered_code(kospi)
        kosdaq = myWindow.filtered_code(kosdaq)

        countTotalStocks = len(kospi) + len(kosdaq)

        # 10%, 5%, 3%, 1%, -1%, -3%, -5%, -10%

        countCodesUp10 = 0
        countCodesUp5 = 0
        countCodesUp3 = 0
        countCodesUp1 = 0
        countCodesDown10 = 0
        countCodesDown5 = 0
        countCodesDown3 = 0
        countCodesDown1 = 0
        countStartUp1 = 0
        countStartUp3 = 0
        countStartUp5 = 0
        countStartUp10 = 0
        countStartDown1 = 0
        countStartDown3 = 0
        countStartDown5 = 0
        countStartDown10 = 0
        countHighUp1 = 0
        countHighUp3 = 0
        countHighUp5 = 0
        countHighUp10 = 0
        countHighDown1 = 0
        countHighDown3 = 0
        countHighDown5 = 0
        countHighDown10 = 0
        countLowUp1 = 0
        countLowUp3 = 0
        countLowUp5 = 0
        countLowUp10 = 0
        countLowDown1 = 0
        countLowDown3 = 0
        countLowDown5 = 0
        countLowDown10 = 0


        # 매일 컬럼명 [종가, 시가, 고가, 저가], 행 [일자] 엑셀 저장
        for condition in conditionList:
            if condition.split('^')[1] == "종가1퍼상승마감":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countCodesUp1 = len(codeList)
            elif condition.split('^')[1] == "종가3퍼상승마감":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countCodesUp3 = len(codeList)
            elif condition.split('^')[1] == "종가5퍼상승마감":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countCodesUp5 = len(codeList)
            elif condition.split('^')[1] == "종가10퍼상승마감":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countCodesUp10 = len(codeList)
            elif condition.split('^')[1] == "종가1퍼하락마감":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countCodesDown1 = len(codeList)
            elif condition.split('^')[1] == "종가3퍼하락마감":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countCodesDown3 = len(codeList)
            elif condition.split('^')[1] == "종가5퍼하락마감":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countCodesDown5 = len(codeList)
            elif condition.split('^')[1] == "종가10퍼하락마감":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countCodesDown10 = len(codeList)
            ######## 시가
            elif condition.split('^')[1] == "시가1퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countStartUp1 = len(codeList)
            elif condition.split('^')[1] == "시가3퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countStartUp3 = len(codeList)
            elif condition.split('^')[1] == "시가5퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countStartUp5 = len(codeList)
            elif condition.split('^')[1] == "시가10퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countStartUp10 = len(codeList)
            elif condition.split('^')[1] == "시가1퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countStartDown1 = len(codeList)
            elif condition.split('^')[1] == "시가3퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countStartDown3 = len(codeList)
            elif condition.split('^')[1] == "시가5퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countStartDown5 = len(codeList)
            elif condition.split('^')[1] == "시가10퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countStartDown10 = len(codeList)
            #### 고가
            elif condition.split('^')[1] == "고가1퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countHighUp1 = len(codeList)
            elif condition.split('^')[1] == "고가3퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countHighUp3 = len(codeList)
            elif condition.split('^')[1] == "고가5퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countHighUp5 = len(codeList)
            elif condition.split('^')[1] == "고가10퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countHighUp10 = len(codeList)
            elif condition.split('^')[1] == "고가1퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countHighDown1 = len(codeList)
            elif condition.split('^')[1] == "고가3퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countHighDown3 = len(codeList)
            elif condition.split('^')[1] == "고가5퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countHighDown5 = len(codeList)
            elif condition.split('^')[1] == "고가10퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countHighDown10 = len(codeList)
            #### 저가
            elif condition.split('^')[1] == "저가1퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countLowUp1 = len(codeList)
            elif condition.split('^')[1] == "저가3퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countLowUp3 = len(codeList)
            elif condition.split('^')[1] == "저가5퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countLowUp5 = len(codeList)
            elif condition.split('^')[1] == "저가10퍼상승":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countLowUp10 = len(codeList)
            elif condition.split('^')[1] == "저가1퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countLowDown1 = len(codeList)
            elif condition.split('^')[1] == "저가3퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countLowDown3 = len(codeList)
            elif condition.split('^')[1] == "저가5퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countLowDown5 = len(codeList)
            elif condition.split('^')[1] == "저가10퍼하락":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
                countLowDown10 = len(codeList)
            else:
                continue

        print("종가1퍼상승마감 = {0}\n종가3퍼상승마감 = {1}\n종가5퍼상승마감 = {2}\n종가10퍼상승마감 = {3}\n"
              "종가1퍼하락마감 = {4}\n종가3퍼하락마감 = {5}\n종가5퍼하락마감 = {6}\n종가10퍼하락마감 = {7}\n전체종목 = {8}"
              .format(countCodesUp1, countCodesUp3, countCodesUp5, countCodesUp10
                      , countCodesDown1, countCodesDown3, countCodesDown5, countCodesDown10, countTotalStocks))

        # 시장현황 엑셀 데이터프레임 저장
        date = myWindow.calculateBusinessDay(0)
        fileName = "C:/Users/A/Desktop/주식자료/MarketCondtion.xlsx"
        sheetName = "Sheet1"
        df = pd.DataFrame([])
        df = myWindow.readExcelToDataFrame(fileName, sheetName)

        # 기존 현황 + 새 현황
        if not df.empty:
            df = df.append(
                {
                    '날짜': date, '총':countTotalStocks, '종1': countCodesUp1, '종3': countCodesUp3, '종5': countCodesUp5, '종10': countCodesUp10
                    , '종-1': countCodesDown1, '종-3': countCodesDown3, '종-5': countCodesDown5, '종-10': countCodesDown10
                    ,'시1': countStartUp1, '시3': countStartUp3, '시5': countStartUp5, '시10': countStartUp10
                    , '시-1': countStartDown1, '시-3': countStartDown3, '시-5': countStartDown5, '시-10':countStartDown10
                    ,'고1': countHighUp1, '고3': countHighUp3, '고5': countHighUp5, '고10': countHighUp10
                    , '고-1': countHighDown1, '고-3': countHighDown3, '고-5': countHighDown5, '고-10':countHighDown10
                    ,'저1': countLowUp1, '저3': countLowUp3, '저5': countLowUp5, '저10': countLowUp10
                    , '저-1': countLowDown1, '저-3': countLowDown3, '저-5': countLowDown5, '저-10':countLowDown10
                }
            , ignore_index=True)
        else:
            df = pd.DataFrame([], columns=['날짜', '총', '종1', '종3', '종5', '종10', '종-1', '종-3', '종-5', '종-10'
                                            , '시1', '시3', '시5', '시10', '시-1', '시-3', '시-5', '시-10'
                                            , '고1', '고3', '고5', '고10', '고-1', '고-3', '고-5', '고-10'
                                            , '저1', '저3', '저5', '저10', '저-1', '저-3', '저-5', '저-10'])
            df = df.append(
                {
                    '날짜': date, '총':countTotalStocks, '종1': countCodesUp1, '종3': countCodesUp3, '종5': countCodesUp5, '종10': countCodesUp10
                    , '종-1': countCodesDown1, '종-3': countCodesDown3, '종-5': countCodesDown5, '종-10': countCodesDown10
                    ,'시1': countStartUp1, '시3': countStartUp3, '시5': countStartUp5, '시10': countStartUp10
                    , '시-1': countStartDown1, '시-3': countStartDown3, '시-5': countStartDown5, '시-10':countStartDown10
                    ,'고1': countHighUp1, '고3': countHighUp3, '고5': countHighUp5, '고10': countHighUp10
                    , '고-1': countHighDown1, '고-3': countHighDown3, '고-5': countHighDown5, '고-10':countHighDown10
                    ,'저1': countLowUp1, '저3': countLowUp3, '저5': countLowUp5, '저10': countLowUp10
                    , '저-1': countLowDown1, '저-3': countLowDown3, '저-5': countLowDown5, '저-10':countLowDown10
                }
            , ignore_index=True)

        # 데이터 프레임 엑셀저장
        df = df.drop_duplicates(subset=['날짜'], keep='last')
        myWindow.saveDataframeToExcel(df, fileName, sheetName)

        # 데이터 프레임 plot -> png 저장
        df = df.apply(pd.to_numeric)
        df['DateTime'] = pd.to_datetime(df['날짜'].astype(str), format='%Y%m%d')
        df['day_of_week'] = df['DateTime'].dt.day_name()

        df["종1"].plot(kind='line', marker='o', color='#0400ff', figsize=(24, 15), fontsize=20)
        df["종-1"].plot(kind='line', marker='o', color='#ff0000', figsize=(24, 15), fontsize=20)
        plot.xticks(range(len(df)), df["day_of_week"].values.tolist(), rotation=90, ha='right', fontsize=15)
        plot.savefig("C:/Users/zzang/Desktop/주식자료/picture/MarketCondtion.png", dpi=100)

        if datetime.datetime.now().hour < 19:
            os.system("shutdown -s -t 60")
    except Exception as e:
        if datetime.datetime.now().hour < 19:
            os.system("shutdown -s -t 60")
        print(e)

    # app.exec_()
