import stockAnalysis
import os
import sys
from PyQt5.QtCore import QEventLoop
from PyQt5.QtWidgets import *
from PyQt5.QAxContainer import *
from PyQt5.QtCore import QEventLoop, QSize
import time
import pandas as pd
import requests
import exchange_calendars as ecals
import datetime
from openpyxl import load_workbook
import random
import numpy as np
import urllib.request
from bs4 import BeautifulSoup
import matplotlib.pyplot as plot

class MyWindow(QMainWindow):
    delay = 1

    def __init__(self):

        super().__init__()

        # 종료 메세지 출력
        self.setMinimumSize(QSize(500, 100))
        self.setWindowTitle("Alert Message")

        pybutton = QPushButton('버튼 눌러야 컴퓨터 계속 사용 가능', self)
        pybutton.clicked.connect(self.clickMethod)
        pybutton.resize(400, 30)
        pybutton.move(50, 35)

        self.kiwoom = QAxWidget("KHOPENAPI.KHOpenAPICtrl.1")
        ### event ###
        self.kiwoom.OnEventConnect.connect(self.event_connect)
        self.kiwoom.OnReceiveTrCondition.connect(self.receive_trCondition)
        self.kiwoom.OnReceiveConditionVer.connect(self.receive_VerCondition)

    def clickMethod(self):
        QMessageBox.about(self, "", "자동종료 취소")
        sys.exit()

    def receive_VerCondition(self, ret, msg):
        self.receive_VerCondition.exit()

    def receive_trCondition(self, screenNo, codes, conditionName, conditionIndex, inquiry):
        print("@@ trCondition 시작")
        if codes == "":
            print("@@ 검색결과 없음")
        self.codeList = codes.split(';')
        del self.codeList[-1]
        self.receive_trCondition_loop.exit()

    def event_connect(self, err_code):
        if err_code == 0:
            print("@@ 로그인 성공")
        self.login_event_loop.exit()

    ### 한국거래소 영업일 계산
    def calculateBusinessDay(self, day):

        global todayDay

        # day = x 일때, x 일전 영업일을 출력
        day = day * -1
        XKRX = ecals.get_calendar("XKRX")  # 한국 코드

        count = 0
        targetDay = 0
        dayCount = day + 2

        for x in range(dayCount - 1):
            while targetDay != x + 1:
                if XKRX.is_session(datetime.date.today() - datetime.timedelta(days=count)):
                    targetDay = targetDay + 1
                count = count + 1
            todayDay = str(datetime.date.today() - datetime.timedelta(days=count - 1)).replace("-", "")

        return todayDay

    ### 실행 함수###
    def kiwoom_login(self):
        print("@@ 로그인 시작")
        self.login_event_loop = QEventLoop()
        self.kiwoom.CommConnect()
        self.login_event_loop.exec_()

    # 조건식 불러오기
    def load_conditionList(self):
        print("@@ 조건식 가져오기 시작")
        self.receive_VerCondition = QEventLoop()

        isLoad = self.kiwoom.GetConditionLoad()
        self.receive_VerCondition.exec_()
        conditions = self.kiwoom.dynamicCall("GetConditionNameList()")
        if isLoad == 1:
            print("@@ 조건식 가져오기 성공")
        else:
            print("@@ 조건식 가져오기 실패")
        time.sleep(self.delay)
        return conditions

    # 조건식 검색, 코드리스트 불러오기
    def search_condition(self, conditionName, conditionIdx):
        print("@@ 조건검색 시작")

        self.receive_trCondition_loop = QEventLoop()
        screenNo = "0150"
        isRealTime = 0
        isRequest = self.kiwoom.dynamicCall("SendCondition(QString, QString, int, int)",
                                            screenNo, conditionName, conditionIdx, isRealTime)
        self.receive_trCondition_loop.exec_()
        if isRequest == 0:
            print("@@ 조건검색 실패")
        if isRequest == 1:
            print("@@ 조건검색 성공")
        time.sleep(self.delay)
        return self.codeList

    # 주식정보 불러오기
    def getStockInfo(self, code, pages):

        df = pd.DataFrame()
        for page in range(1, pages):
            url = "https://finance.naver.com/item/sise_day.naver?code={}&page={}".format(code, page)
            res = requests.get(url, headers={'User-agent': 'Mozilla/5.0'})
            df = pd.concat([df, pd.read_html(res.text, header=0)[0]], axis=0)
            df = df.reset_index(drop=True)
            # df['저가-시가'] = df['저가'] - df['시가']
            # df['고가-시가'] = df['고가'] - df['시가']

        # df.dropna()를 이용해 결측값 있는 행 제거
        df = df.dropna()
        return df

    # 전체 종목코드 가져오기
    def GetCodeListByMarket(self, market):
        codeList = self.kiwoom.dynamicCall("GetCodeListByMarket(QString)", market)
        return codeList

    # 종목코드로 종목명 불러오기
    def GetMasterCodeName(self, code):
        code_name = self.kiwoom.dynamicCall("GetMasterCodeName(QString)", code)
        return code_name

    # 종목 매수 (in_strAccount, in_strCode, out_result)

    # 종목 매도 (in_strAccount, in_strCode, out_result)

    # 보유 종목 출력 (in_strAccount, out_listCodes)

    # 주식정보 엑셀 저장하기
    def saveDataframeToExcel(self, df, fileName, sheetName):
        if os.path.isfile(fileName):
            ExcelWorkbook = load_workbook(fileName)
            if sheetName in ExcelWorkbook.sheetnames:
                ExcelWorkbook.remove(ExcelWorkbook[sheetName])
            # Generating the writer engine
            writer = pd.ExcelWriter(fileName, engine='openpyxl')
            # Assigning the workbook to the writer engine
            writer.book = ExcelWorkbook
            df.to_excel(writer, index=False, sheet_name=sheetName)
            writer.save()
            writer.close()
        else:
            df.to_excel(fileName, index=False, sheet_name=sheetName)
        # 주식정보 비교하기

    # 엑셀불러오기
    def readExcelToDataFrame(self, fileName, sheetName):
        global df
        if os.path.isfile(fileName):
            df = pd.read_excel(fileName, dtype=str, sheet_name=sheetName)
        else:
            pass
        return df

    def compare_stockInfo(self, code):
        pass

    # 특정 시트에 종목 넣기
    def insert_StockToSheet(self, code, sheet):
        pass

    # 날짜별 수익률 계산 (인수 : DF)
    def ReturnMoneyPerDay(self, df):
        dfDate = df.copy().drop_duplicates(subset=['날짜'], keep='last')
        lstDate = []
        dfResult = pd.DataFrame([], columns=['날짜3', '총수익률', '수익종목수', '손해종목수'])
        for index, row in dfDate.iterrows():
            lstDate.append(row['날짜'])
        for i in lstDate:
            # 해당 날짜 분류
            dfSort = df.loc[df['날짜'] == i].copy()
            # 결측값 제거
            dfSort['수익률'].replace('', np.nan, inplace=True)
            dfSort = dfSort.dropna(subset=['수익률'])
            dfSort['수익률'] = dfSort['수익률'].astype(int)
            # dfSort = dfSort['수익률']
            Total = dfSort['수익률'].sum()
            benefit = dfSort.loc[dfSort['수익률'] == 1, '수익률'].count()
            loss = dfSort.loc[dfSort['수익률'] == -5, '수익률'].count()

            newRow = {'날짜3': [i],
                      '총수익률': [Total],
                      '수익종목수': [benefit],
                      '손해종목수': [loss]}
            newResult = pd.DataFrame(data=newRow)
            dfResult = pd.concat([dfResult, newResult], ignore_index=True)
        KOSPIUpDown = myWindow.getKOSPIInfo()
        KOSDAQUpDown = myWindow.getKOSDAQInfo()
        newRow = {'날짜3': ['KOSPI'],
                  '총수익률': [KOSPIUpDown],
                  '수익종목수': ['KOSDAQ'],
                  '손해종목수': [KOSDAQUpDown]}
        newResult = pd.DataFrame(data=newRow)
        dfResult = pd.concat([dfResult, newResult], ignore_index=True)
        return dfResult

    # 종료 메세지 박스 띄우기
    def quitDialog(self):
        pass

    # 코드리스트 필터링
    def filtered_code(self, list_code):
        new_code_list = []
        for code in list_code:
            codeName = self.GetMasterCodeName(code)
            # 스팩주 제거
            if "스팩" in codeName or "4호" in codeName:
                pass
            elif "ETN" in codeName:
                pass
            elif codeName[-1:] == "우" or "3우C" in codeName or "우B" in codeName or "G3우" in codeName:
                pass
            elif "TIGER" in codeName or "KOSEF" in codeName or "KBSTAR" in codeName or "KODEX" in codeName or "KINDEX" in codeName or "TREX" in codeName:
                pass
            elif "ARIRANG" in codeName or "SMART" in codeName or "FOCUS" in codeName or "HANARO" in codeName or "TIMEFOLIO" in codeName or "네비게이터" in codeName:
                pass
            else:
                new_code_list.append(code)

        return new_code_list


if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        myWindow = MyWindow()
        conditionName = ""
        processStage = ""

        myWindow.show()
        # 키움 로그인
        myWindow.kiwoom_login()

        # 조건식 불러오기
        conditionList = myWindow.load_conditionList().split(";")
        del conditionList[-1]
        conditionListCopy = conditionList
        conditionList = []

        # 조건식 이름 변경
        processStage = "조건식 이름변경"
        for name in conditionListCopy:
            name = name.replace("/", "")
            conditionList.append(name)

        # 매일 컬럼명 [종가, 시가, 고가, 저가], 행 [일자] 엑셀 저장
        for condition in conditionList:
            if condition.split('^')[1] == "종가1퍼상승마감":
                codeList = myWindow.search_condition(condition.split('^')[1], condition.split('^')[0])
                codeList = myWindow.filtered_code(codeList)
            else:
                continue

        # 시장현황 엑셀 데이터프레임 저장
        processStage = "df생성"
        date = myWindow.calculateBusinessDay(0)
        fileName = "C:/Users/zzang/Desktop/주식자료/종가1이상20일자료.xlsx"
        sheetName = date
        df = pd.DataFrame([])
        total_df = pd.DataFrame([], columns=['코드명', '종증감평균', '종증감0', '거증감0', '종가0', '고가0', '저가0', '시가0', '거래량0'
                                        , '종증감1', '거증감1', '종가1', '고가1', '저가1', '시가1', '거래량1'
                                        , '종증감2', '거증감2', '종가2', '고가2', '저가2', '시가2', '거래량2'
                                        , '종증감3', '거증감3', '종가3', '고가3', '저가3', '시가3', '거래량3'
                                        , '종증감4', '거증감4', '종가4', '고가4', '저가4', '시가4', '거래량4'
                                        , '종증감5', '거증감5', '종가5', '고가5', '저가5', '시가5', '거래량5'
                                        , '종증감6', '거증감6', '종가6', '고가6', '저가6', '시가6', '거래량6'
                                        , '종증감7', '거증감7', '종가7', '고가7', '저가7', '시가7', '거래량7'
                                        , '종증감8', '거증감8', '종가8', '고가8', '저가8', '시가8', '거래량8'
                                        , '종증감9', '거증감9', '종가9', '고가9', '저가9', '시가9', '거래량9'
                                        , '종증감10', '거증감10', '종가10', '고가10', '저가10', '시가10', '거래량10'
                                        , '종증감11', '거증감11', '종가11', '고가11', '저가11', '시가11', '거래량11'
                                        , '종증감12', '거증감12', '종가12', '고가12', '저가12', '시가12', '거래량12'
                                        , '종증감13', '거증감13', '종가13', '고가13', '저가13', '시가13', '거래량13'
                                        , '종증감14', '거증감14', '종가14', '고가14', '저가14', '시가14', '거래량14'
                                        , '종증감15', '거증감15', '종가15', '고가15', '저가15', '시가15', '거래량15'
                                        , '종증감16', '거증감16', '종가16', '고가16', '저가16', '시가16', '거래량16'
                                        , '종증감17', '거증감17', '종가17', '고가17', '저가17', '시가17', '거래량17'
                                        , '종증감18', '거증감18', '종가18', '고가18', '저가18', '시가18', '거래량18'
                                        , '종증감19', '거증감19', '종가19', '고가19', '저가19', '시가19', '거래량19'])

        processStage = "코드 반복"
        try:
            for idx, code in enumerate(codeList):
                print("완료 {0} / {1} ::: {2}".format(idx+1, len(codeList), code))

                # 주식정보 불러오기 (네이버 크롤링)
                df = myWindow.getStockInfo(code, 3)

                if len(df.index) < 20:
                    continue

                # 새로운 정보 DF 생성
                new_df = {'코드명': code, '종증감평균': 0
                        , '종가0': df.at[1, '종가'], '고가0': df.at[1, '고가'], '저가0': df.at[1, '저가'], '시가0': df.at[1, '시가'],'거래량0': df.at[1, '거래량']
                        , '종가1': df.at[2, '종가'], '고가1': df.at[2, '고가'], '저가1': df.at[2, '저가'],'시가1': df.at[2, '시가'],'거래량1': df.at[2, '거래량']
                        , '종가2': df.at[3, '종가'], '고가2': df.at[3, '고가'], '저가2': df.at[3, '저가'],'시가2': df.at[3, '시가'],'거래량2': df.at[3, '거래량']
                        , '종가3': df.at[4, '종가'], '고가3': df.at[4, '고가'], '저가3': df.at[4, '저가'],'시가3': df.at[4, '시가'],'거래량3': df.at[4, '거래량']
                        , '종가4': df.at[5, '종가'], '고가4': df.at[5, '고가'], '저가4': df.at[5, '저가'],'시가4': df.at[5, '시가'],'거래량4': df.at[5, '거래량']
                        , '종가5': df.at[9, '종가'], '고가5': df.at[9, '고가'], '저가5': df.at[9, '저가'],'시가5': df.at[9, '시가'],'거래량5': df.at[9, '거래량']
                        , '종가6': df.at[10, '종가'], '고가6': df.at[10, '고가'], '저가6': df.at[10, '저가'],'시가6': df.at[10, '시가'],'거래량6': df.at[10, '거래량']
                        , '종가7': df.at[11, '종가'], '고가7': df.at[11, '고가'], '저가7': df.at[11, '저가'],'시가7': df.at[11, '시가'],'거래량7': df.at[11, '거래량']
                        , '종가8': df.at[12, '종가'], '고가8': df.at[12, '고가'], '저가8': df.at[12, '저가'],'시가8': df.at[12, '시가'],'거래량8': df.at[12, '거래량']
                        , '종가9': df.at[13, '종가'], '고가9': df.at[13, '고가'], '저가9': df.at[13, '저가'],'시가9': df.at[13, '시가'],'거래량9': df.at[13, '거래량']
                        , '종가10': df.at[16, '종가'], '고가10': df.at[16, '고가'], '저가10': df.at[16, '저가'],'시가10': df.at[16, '시가'],'거래량10': df.at[16, '거래량']
                        , '종가11': df.at[17, '종가'], '고가11': df.at[17, '고가'], '저가11': df.at[17, '저가'],'시가11': df.at[17, '시가'],'거래량11': df.at[17, '거래량']
                        , '종가12': df.at[18, '종가'], '고가12': df.at[18, '고가'], '저가12': df.at[18, '저가'],'시가12': df.at[18, '시가'],'거래량12': df.at[18, '거래량']
                        , '종가13': df.at[19, '종가'], '고가13': df.at[19, '고가'], '저가13': df.at[19, '저가'],'시가13': df.at[19, '시가'],'거래량13': df.at[19, '거래량']
                        , '종가14': df.at[20, '종가'], '고가14': df.at[20, '고가'], '저가14': df.at[20, '저가'],'시가14': df.at[20, '시가'],'거래량14': df.at[20, '거래량']
                        , '종가15': df.at[24, '종가'], '고가15': df.at[24, '고가'], '저가15': df.at[24, '저가'],'시가15': df.at[24, '시가'], '거래량15': df.at[24, '거래량']
                        , '종가16': df.at[25, '종가'], '고가16': df.at[25, '고가'], '저가16': df.at[25, '저가'],'시가16': df.at[25, '시가'], '거래량16': df.at[25, '거래량']
                        , '종가17': df.at[26, '종가'], '고가17': df.at[26, '고가'], '저가17': df.at[26, '저가'],'시가17': df.at[26, '시가'], '거래량17': df.at[26, '거래량']
                        , '종가18': df.at[27, '종가'], '고가18': df.at[27, '고가'], '저가18': df.at[27, '저가'],'시가18': df.at[27, '시가'], '거래량18': df.at[27, '거래량']
                        , '종가19': df.at[28, '종가'], '고가19': df.at[28, '고가'], '저가19': df.at[28, '저가'],'시가19': df.at[28, '시가'], '거래량19': df.at[28, '거래량']
                        , '종증감0': round(df.at[1, '종가'] / df.at[2, '종가'], 3),'거증감0': round(df.at[1, '거래량'] / df.at[2, '거래량'], 3)
                        , '종증감1': round(df.at[2, '종가'] / df.at[3, '종가'], 3), '거증감1': round(df.at[2, '거래량'] / df.at[3, '거래량'], 3)
                        , '종증감2': round(df.at[3, '종가'] / df.at[4, '종가'], 3), '거증감2': round(df.at[3, '거래량'] / df.at[4, '거래량'], 3)
                        , '종증감3': round(df.at[4, '종가'] / df.at[5, '종가'], 3), '거증감3': round(df.at[4, '거래량'] / df.at[5, '거래량'], 3)
                        , '종증감4': round(df.at[5, '종가'] / df.at[9, '종가'], 3), '거증감4': round(df.at[5, '거래량'] / df.at[9, '거래량'], 3)
                        , '종증감5': round(df.at[9, '종가'] / df.at[10, '종가'], 3), '거증감5': round(df.at[9, '거래량'] / df.at[10, '거래량'], 3)
                        , '종증감6': round(df.at[10, '종가'] / df.at[11, '종가'], 3), '거증감6': round(df.at[10, '거래량'] / df.at[11, '거래량'], 3)
                        , '종증감7': round(df.at[11, '종가'] / df.at[12, '종가'], 3), '거증감7': round(df.at[11, '거래량'] / df.at[12, '거래량'], 3)
                        , '종증감8': round(df.at[12, '종가'] / df.at[13, '종가'], 3), '거증감8': round(df.at[12, '거래량'] / df.at[13, '거래량'], 3)
                        , '종증감9': round(df.at[13, '종가'] / df.at[16, '종가'], 3), '거증감9': round(df.at[13, '거래량'] / df.at[16, '거래량'], 3)
                        , '종증감10': round(df.at[16, '종가'] / df.at[17, '종가'], 3), '거증감10': round(df.at[16, '거래량'] / df.at[17, '거래량'], 3)
                        , '종증감11': round(df.at[17, '종가'] / df.at[18, '종가'], 3), '거증감11': round(df.at[17, '거래량'] / df.at[18, '거래량'], 3)
                        , '종증감12': round(df.at[18, '종가'] / df.at[19, '종가'], 3), '거증감12': round(df.at[18, '거래량'] / df.at[19, '거래량'], 3)
                        , '종증감13': round(df.at[19, '종가'] / df.at[20, '종가'], 3), '거증감13': round(df.at[19, '거래량'] / df.at[20, '거래량'], 3)
                        , '종증감14': round(df.at[20, '종가'] / df.at[24, '종가'], 3), '거증감14': round(df.at[20, '거래량'] / df.at[24, '거래량'], 3)
                        , '종증감15': round(df.at[24, '종가'] / df.at[25, '종가'], 3), '거증감15': round(df.at[24, '거래량'] / df.at[25, '거래량'], 3)
                        , '종증감16': round(df.at[25, '종가'] / df.at[26, '종가'], 3), '거증감16': round(df.at[25, '거래량'] / df.at[26, '거래량'], 3)
                        , '종증감17': round(df.at[26, '종가'] / df.at[27, '종가'], 3), '거증감17': round(df.at[26, '거래량'] / df.at[27, '거래량'], 3)
                        , '종증감18': round(df.at[27, '종가'] / df.at[28, '종가'], 3), '거증감18': round(df.at[27, '거래량'] / df.at[28, '거래량'], 3)
                    }

                new_df['종증감평균'] = round((new_df['종증감1'] + new_df['종증감2'] + new_df['종증감3'] + new_df['종증감4'] + \
                                  new_df['종증감5'] + new_df['종증감6'] + new_df['종증감7'] + new_df['종증감8'] + \
                                  new_df['종증감9'] + new_df['종증감10'] + new_df['종증감11'] + new_df['종증감12'] + \
                                  new_df['종증감13'] + new_df['종증감14'] + new_df['종증감15'] + new_df['종증감16'] + \
                                  new_df['종증감17'] + new_df['종증감18'] + new_df['종증감0']) / 19, 3)

                # 새 데이터 프레임에 주식정보 저장

                total_df = total_df.append(
                    new_df, ignore_index=True)
                # 새 데이터 정보 이미지 저장
                path = os.path.join('C:/Users/zzang/Desktop/주식자료/picture/', date)
                if idx == 0:
                    if not os.path.exists(path):
                        os.makedirs(path)

                img_data = {'period': [18, 17, 16, 15, 14, 13, 12, 11, 10, 9, 8, 7, 6, 5, 4, 3, 2, 1, 0],
                        'rate': [new_df['종증감0'], new_df['종증감1'], new_df['종증감2'], new_df['종증감3'], new_df['종증감4'], new_df['종증감5'],
                                 new_df['종증감6'], new_df['종증감7'], new_df['종증감8'], new_df['종증감9'], new_df['종증감10'],
                                 new_df['종증감11'], new_df['종증감12'], new_df['종증감13'], new_df['종증감14'], new_df['종증감15'],
                                 new_df['종증감16'], new_df['종증감17'], new_df['종증감18']]}

                # 평균
                aver_rate = (new_df['종증감0'] + new_df['종증감1'] + new_df['종증감2'] + new_df['종증감3'] + new_df['종증감4'] + \
                new_df['종증감5'] + new_df['종증감6'] + new_df['종증감7'] + new_df['종증감8'] + new_df['종증감9'] + \
                new_df['종증감10'] +   new_df['종증감11'] + new_df['종증감12'] + new_df['종증감13'] + new_df['종증감14'] + \
                new_df['종증감15'] + new_df['종증감16'] + new_df['종증감17'] + new_df['종증감18']) / 19

                img_df = pd.DataFrame(img_data)
                img_df.plot(x='period', y='rate', kind='line', figsize=(10, 8), fontsize=20, marker='o', color="black")
                plot.axhline(y=1, color='r', linestyle='dashed')
                plot.axhline(y=aver_rate, color='g', linestyle='dashed')

                # MarketCondition 불러오기
                df_marketCondtion = myWindow.readExcelToDataFrame("C:/Users/zzang/Desktop/주식자료/MarketCondtion.xlsx", "Sheet1")


                # plot 저장
                save_path = path + '/' + code + '.png'
                plot.savefig(save_path, dpi=50)
                plot.close()

        except Exception as e:
            print(e)

        # 데이터 프레임 엑셀저장
        # df = df.drop_duplicates(subset=['날짜'], keep='last')
        myWindow.saveDataframeToExcel(total_df, fileName, sheetName)
        if datetime.datetime.now().hour < 19:
            #os.system("shutdown -s -t 60")
            pass
    except Exception as e:
        if datetime.datetime.now().hour < 19:
            #os.system("shutdown -s -t 60")
            pass
        print(processStage)
        print(e)

    # app.exec_()
