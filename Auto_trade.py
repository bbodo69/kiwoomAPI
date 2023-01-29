import stockAnalysis
import os
import sys
from PyQt5.QtCore import QEventLoop
from PyQt5.QtWidgets import *
from PyQt5.QAxContainer import *
from PyQt5.QtCore import QEventLoop, QSize
import time
import pandas as pd
import requests
import exchange_calendars as ecals
import datetime
from openpyxl import load_workbook
import random
import numpy as np
import urllib.request
from bs4 import BeautifulSoup
import matplotlib.pyplot as plot


class MyWindow(QMainWindow):
    delay = 3.6

    def __init__(self):

        ####### 계좌 관련 변수
        self.account_num = None
        self.deposit = 0
        self.use_money = 0
        self.use_money_percent = 0
        self.output_deposit = 0
        self.account_list = []
        self.account_stock_dict = {}    # 보유종목

        super().__init__()

        # 종료 메세지 출력
        self.setMinimumSize(QSize(500, 100))
        self.setWindowTitle("Alert Message")

        pybutton = QPushButton('버튼 눌러야 컴퓨터 계속 사용 가능', self)
        pybutton.clicked.connect(self.clickMethod)
        pybutton.resize(400, 30)
        pybutton.move(50, 35)

        self.kiwoom = QAxWidget("KHOPENAPI.KHOpenAPICtrl.1")
        ### event ###
        self.kiwoom.OnEventConnect.connect(self.event_connect)
        self.kiwoom.OnReceiveTrCondition.connect(self.receive_trCondition)
        self.kiwoom.OnReceiveConditionVer.connect(self.receive_VerCondition)
        self.kiwoom.OnReceiveChejanData.connect(self.receive_Chejan)
        self.kiwoom.OnReceiveTrData.connect(self.receive_trdata)

    def receive_Chejan(self, sGubun,nItemCnt, sFidlist):
        print(sGubun, nItemCnt, sFidlist)

    def clickMethod(self):
        QMessageBox.about(self, "", "자동종료 취소")
        sys.exit()

    def receive_VerCondition(self, ret, msg):
        self.receive_VerCondition.exit()

    def receive_trCondition(self, screenNo, codes, conditionName, conditionIndex, inquiry):
        print("@@ trCondition 시작")
        if codes == "":
            print("@@ 검색결과 없음")
        self.codeList = codes.split(';')
        del self.codeList[-1]
        self.receive_trCondition_loop.exit()

    def event_connect(self, err_code):
        if err_code == 0:
            print("@@ 로그인 성공")
        self.login_event_loop.exit()

    ### 한국거래소 영업일 계산
    def calculateBusinessDay(self, day):
        # day = x 일때, x 일전 영업일을 출력
        day = day * -1
        XKRX = ecals.get_calendar("XKRX")  # 한국 코드

        count = 0
        targetDay = 0
        dayCount = day + 2

        for x in range(dayCount - 1):
            while targetDay != x + 1:
                if XKRX.is_session(datetime.date.today() - datetime.timedelta(days=count)):
                    targetDay = targetDay + 1
                count = count + 1
            todayDay = str(datetime.date.today() - datetime.timedelta(days=count - 1)).replace("-", "")

        return todayDay

    ### 실행 함수###
    def kiwoom_login(self):
        print("@@ 로그인 시작")
        self.login_event_loop = QEventLoop()
        self.kiwoom.CommConnect()
        self.login_event_loop.exec_()

    # 조건식 불러오기
    def load_conditionList(self):
        print("@@ 조건식 가져오기 시작")
        self.receive_VerCondition = QEventLoop()

        isLoad = self.kiwoom.GetConditionLoad()
        self.receive_VerCondition.exec_()
        conditions = self.kiwoom.dynamicCall("GetConditionNameList()")
        if isLoad == 1:
            print("@@ 조건식 가져오기 성공")
        else:
            print("@@ 조건식 가져오기 실패")
        time.sleep(self.delay)
        return conditions

    # 조건식 검색, 코드리스트 불러오기
    def search_condition(self, conditionName, conditionIdx):
        print("@@ 조건검색 시작")

        self.receive_trCondition_loop = QEventLoop()
        screenNo = "0150"
        isRealTime = 0
        isRequest = self.kiwoom.dynamicCall("SendCondition(QString, QString, int, int)",
                                            screenNo, conditionName, conditionIdx, isRealTime)
        self.receive_trCondition_loop.exec_()
        if isRequest == 0:
            print("@@ 조건검색 실패")
        if isRequest == 1:
            print("@@ 조건검색 성공")
        time.sleep(self.delay)
        return self.codeList

    # 주식정보 불러오기
    def getStockInfo(self, code, pages):
        df = pd.DataFrame()
        for page in range(1, pages):
            url = "https://finance.naver.com/item/sise_day.naver?code={}&page={}".format(code, page)
            res = requests.get(url, headers={'User-agent': 'Mozilla/5.0'})
            df = pd.concat([df, pd.read_html(res.text, header=0)[0]], axis=0)
            # df['저가-시가'] = df['저가'] - df['시가']
            # df['고가-시가'] = df['고가'] - df['시가']

        # df.dropna()를 이용해 결측값 있는 행 제거
        df = df.dropna()
        return df

    # 전체 종목코드 가져오기
    def GetCodeListByMarket(self, market):
        codeList = self.kiwoom.dynamicCall("GetCodeListByMarket(QString)", market)
        return codeList

    # 종목코드로 종목명 불러오기
    def GetMasterCodeName(self, code):
        code_name = self.kiwoom.dynamicCall("GetMasterCodeName(QString)", code)
        return code_name

    # 종목 매수 (in_strAccount, in_strCode, out_result)
    def buy_Stock(self, code, amount, accountNumber):
        print("@@ 주식매수")
        self.kiwoom.SendOrder('지정가매수', '0101', accountNumber, 1, code, amount, 0, '03', '')

    # 종목 매도 (in_strAccount, in_strCode, out_result)
    def sell_Stock(self, code, amount, accountNumber):
        print("@@ 주식매도")
        self.kiwoom.SendOrder('지정가매수', '0101', accountNumber, 2, code, amount, 0, '03', '')

    # 보유 종목 출력 (in_strAccount, out_listCodes)

    

    # 주식정보 엑셀 저장하기
    def saveDataframeToExcel(self, df, fileName, sheetName):
        if os.path.isfile(fileName):
            ExcelWorkbook = load_workbook(fileName)
            if sheetName in ExcelWorkbook.sheetnames:
                ExcelWorkbook.remove(ExcelWorkbook[sheetName])
            # Generating the writer engine
            writer = pd.ExcelWriter(fileName, engine='openpyxl')
            # Assigning the workbook to the writer engine
            writer.book = ExcelWorkbook
            df.to_excel(writer, index=False, sheet_name=sheetName)
            writer.save()
            writer.close()
        else:
            df.to_excel(fileName, index=False, sheet_name=sheetName)
        # 주식정보 비교하기

    # 엑셀불러오기
    def readExcelToDataFrame(self, fileName, sheetName):
        df = pd.DataFrame([])
        if os.path.isfile(fileName):
            df = pd.read_excel(fileName, dtype=str, sheet_name=sheetName)
        else:
            pass
        return df

    def compare_stockInfo(self, code):
        pass

    # 특정 시트에 종목 넣기
    def insert_StockToSheet(self, code, sheet):
        pass

    # 날짜별 수익률 계산 (인수 : DF)
    def ReturnMoneyPerDay(self, df):
        dfDate = df.copy().drop_duplicates(subset=['날짜'], keep='last')
        lstDate = []
        dfResult = pd.DataFrame([], columns=['날짜3', '총수익률', '수익종목수', '손해종목수'])
        for index, row in dfDate.iterrows():
            lstDate.append(row['날짜'])
        for i in lstDate:
            # 해당 날짜 분류
            dfSort = df.loc[df['날짜'] == i].copy()
            # 결측값 제거
            dfSort['수익률'].replace('', np.nan, inplace=True)
            dfSort = dfSort.dropna(subset=['수익률'])
            dfSort['수익률'] = dfSort['수익률'].astype(int)
            # dfSort = dfSort['수익률']
            Total = dfSort['수익률'].sum()
            benefit = dfSort.loc[dfSort['수익률'] == 1, '수익률'].count()
            loss = dfSort.loc[dfSort['수익률'] == -5, '수익률'].count()

            newRow = {'날짜3': [i],
                      '총수익률': [Total],
                      '수익종목수': [benefit],
                      '손해종목수': [loss]}
            newResult = pd.DataFrame(data=newRow)
            dfResult = pd.concat([dfResult, newResult], ignore_index=True)
        KOSPIUpDown = myWindow.getKOSPIInfo()
        KOSDAQUpDown = myWindow.getKOSDAQInfo()
        newRow = {'날짜3': ['KOSPI'],
                  '총수익률': [KOSPIUpDown],
                  '수익종목수': ['KOSDAQ'],
                  '손해종목수': [KOSDAQUpDown]}
        newResult = pd.DataFrame(data=newRow)
        dfResult = pd.concat([dfResult, newResult], ignore_index=True)
        return dfResult


    # 종료 메세지 박스 띄우기
    def quitDialog(self):
        pass

    # 코드리스트 필터링
    def filtered_code(self, list_code):
        new_code_list = []
        for code in list_code:
            codeName = self.GetMasterCodeName(code)
            # 스팩주 제거
            if "스팩" in codeName or "4호" in codeName:
                pass
            elif "ETN" in codeName:
                pass
            elif codeName[-1:] == "우" or "3우C" in codeName or "우B" in codeName or "G3우" in codeName:
                pass
            elif "TIGER" in codeName or "KOSEF" in codeName or "KBSTAR" in codeName or "KODEX" in codeName or "KINDEX" in codeName or "TREX" in codeName:
                pass
            elif "ARIRANG" in codeName or "SMART" in codeName or "FOCUS" in codeName or "HANARO" in codeName or "TIMEFOLIO" in codeName or "네비게이터" in codeName:
                pass
            else:
                new_code_list.append(code)

        return new_code_list

    ''''''''''''''''receiveTR'''''''''''''''''

    def receive_trdata(self, screen_no, rqname, trcode, recordname, sPrevNext, data_len, err_code, msg1, msg2):

        print("요청이름 : " + rqname)

        # 예수금
        if rqname == "opw00001_req":
            print("rqName = opw00001_req")
            deposit = self.kiwoom.dynamicCall("GetCommData(QString, QString, int, QString)", trcode, rqname, 0, "예수금")
            self.deposit = int(deposit)
            output_deposit = self.kiwoom.dynamicCall("GetCommData(QString, QString, int, QString)", trcode, rqname, 0, "출금가능금액")
            self.output_deposit = int(output_deposit)

        # 총매입금액
        if rqname == "opw00018_req":
            print("rqName = opw00018_req")
            total_buy_money = self.kiwoom.dynamicCall("GetCommData(QString, QString, int, QString", trcode, rqname, 0, "총매입금액")
            self.total_buy_money = int(total_buy_money)
            # 보유 종목 가져오기
            rows = self.kiwoom.dynamicCall("GetRepeatCnt(QString, QString)", trcode, rqname)
            # 계좌가 보유중인 종목의 갯수를 카운트해준다.
            print("보유종목수 : {0}".format(rows))
            for i in range(rows):
                code = self.kiwoom.dynamicCall("GetCommData(QString, QString, int, QString)",
                                        trcode, rqname, i, "종목번호")
                code = code.strip()[1:]
                code_nm = self.kiwoom.dynamicCall("GetCommData(QString, QString, int, QString)",
                                           trcode, rqname, i, "종목명")
                stock_quantity = self.kiwoom.dynamicCall("GetCommData(QString, QString, int, QString)",
                                                  trcode, rqname, i, "보유수량")
                buy_price = self.kiwoom.dynamicCall("GetCommData(QString, QString, int, QString)",
                                             trcode, rqname, i, "매입가")  # 매입가 평균
                learn_rate = self.kiwoom.dynamicCall("GetCommData(QString, QString, int, QString)",
                                              trcode, rqname, i, "수익률(%)")
                current_price = self.kiwoom.dynamicCall("GetCommData(QString, QString, int, QString)",
                                                 trcode, rqname, i, "현재가")
                total_chegual_price = self.kiwoom.dynamicCall("GetCommData(QString, QString, int, QString)",
                                                       trcode, rqname, i, "매입금액")
                possible_quantity = self.kiwoom.dynamicCall("GetCommData(QString, QString, int, QString)",
                                                     trcode, rqname, i, "매매가능수량")
                if code in self.account_stock_dict:
                    pass
                else:
                    self.account_stock_dict[code] = {}
                code_nm = code_nm.strip()
                stock_quantity = int(stock_quantity.strip())
                buy_price = int(buy_price.strip())
                learn_rate = float(learn_rate.strip())
                current_price = int(current_price.strip())
                total_chegual_price = int(total_chegual_price.strip())
                possible_quantity = int(possible_quantity.strip())

                self.account_stock_dict[code].update({"종목명": code_nm})
                self.account_stock_dict[code].update({"보유수량": stock_quantity})
                self.account_stock_dict[code].update({"매입가": buy_price})
                self.account_stock_dict[code].update({"수익률(%)": learn_rate})
                self.account_stock_dict[code].update({"현재가": current_price})
                self.account_stock_dict[code].update({"매입금액": total_chegual_price})
                self.account_stock_dict[code].update({'매매가능수량': possible_quantity})

                print("종목코드: %s - 종목명: %s - 보유수량: %s - 매입가:%s - 수익률: %s - 현재가: %s" % (
                    code, code_nm, stock_quantity, buy_price, learn_rate, current_price))
                # print(self.account_stock_dict.get('005930')['종목명']), # 가져올때
                #                print("sPrevNext : %s" % sPrevNext)
                #                print("계좌에 가지고 있는 종목은 %s " % rows)
                if sPrevNext == "2":
                    self.detail_account_mystock(sPrevNext="2")
                else:
                    self.detail_account_mystock_loop.exit()

        if rqname == "opt10086_pre_req":
            self.preStart = self.kiwoom.dynamicCAll("CommGetData(QString, QString, QString, int, QString)",
                                                    trcode, "",
                                                    rqname, 0, "시가")
            self.preHigh = self.kiwoom.dynamicCall("CommGetData(QString, QString, QString, int, QString)",
                                                   trcode, "",
                                                   rqname, 0, "고가")
            self.preLow = self.kiwoom.dynamicCall("CommGetData(QString, QString, QString, int, QString)",
                                                  trcode, "",
                                                  rqname, 0, "저가")
            self.preEnd = self.kiwoom.dynamicCall("CommGetData(QString, QString, QString, int, QString)",
                                                  trcode, "",
                                                  rqname, 0, "종가")
            self.preDay = self.kiwoom.dynamicCall("CommGetData(QString, QString, QString, int, QString)",
                                                  trcode, "",
                                                  rqname, 0, "전일비")
            self.preUpDown = self.kiwoom.dynamicCall("CommGetData(QString, QString, QString, int, QString)",
                                                     trcode, "",
                                                     rqname, 0, "등락률")
            # self.opt10086_pre_req_loop.exit()
            print("opt10086_pre_req 끝")

    def get_account_info(self):
        print("get_account_info 시작")
        account_list = self.kiwoom.dynamicCall("GetLoginInfo(QString)", "ACCNO")
        self.account_list = account_list
        return account_list

    def detail_account_mystock(self, account, sPrevNext = "0"):
        self.detail_account_mystock_loop = QEventLoop()
        print("detail_account_mystock 시작")
        self.kiwoom.dynamicCall("SetInputValue(QString, QString)", "계좌번호", account)
        self.kiwoom.dynamicCall("SetInputValue(QString, QString)", "비밀번호", "0000")
        self.kiwoom.dynamicCall("SetInputValue(QString, QString)", "비밀번호입력매체구분", "구분")
        self.kiwoom.dynamicCall("SetInputValue(QString, QString)", "조회구분", "1")
        self.kiwoom.dynamicCall("CommRqData(QString, QString, int, QString)", "opw00018_req", "opw00018", sPrevNext, "0111")
        # CommRqData

        self.detail_account_mystock_loop.exec()
        print("detail_account_mystock 종료")

    def detail_account_info(self, account, sPrevNext = "0"):
        print("detail_account_info 시작")
        self.kiwoom.dynamicCall("SetInputValue(QString, QString)", "계좌번호", account)
        self.kiwoom.dynamicCall("SetInputValue(QString, QString)", "비밀번호", "0000")
        self.kiwoom.dynamicCall("SetInputValue(QString, QString)", "비밀번호입력매체구분", "구분")
        self.kiwoom.dynamicCall("SetInputValue(QString, QString)", "조회구분", "1")
        self.kiwoom.dynamicCall("CommRqData(QString, QString, int, QString)", "opw00001_req", "opw00001", sPrevNext, "0112")

if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        myWindow = MyWindow()


        # 변수저장
        account_num = "8038151811"
        buy_price = 500000
        file_names = ["C:/Users/A/Desktop/주식자료/성장주.xlsx","C:/Users/A/Desktop/주식자료/수익성이 좋은 기업.xlsx"]
        sheet_name = myWindow.calculateBusinessDay(-1)
        df_targetCodesInfo = pd.DataFrame([])
        conditionName = ""
        processStage = ""
        myWindow.show()
        # 키움 로그인
        # myWindow.kiwoom_login()

        # # 조건식 불러오기
        # conditionList = myWindow.load_conditionList().split(";")
        # del conditionList[-1]
        # conditionListCopy = conditionList
        # conditionList = []
        #
        # # 사용할 조건식 선택
        # use_condition_list = ['수익성이 좋은 기업', '가치주', '성장주']
        #
        # # 조건식 이름변경 저장
        # processStage = "조건식 이름변경"
        # for name in conditionListCopy:
        #     if str(name).split("^")[1] in use_condition_list:
        #         name = name.replace("/", "")
        #         conditionList.append(name)
        #
        # # 조건식 코드리스트 저장
        # use_code_list = []
        #
        # for i in conditionList:
        #
        #     conditionName = str(i).split("^")[1]
        #     conditionIndex = str(i).split("^")[0]
        #
        #     use_code_list = use_code_list + myWindow.search_condition(conditionName, conditionIndex)

        # 매수 코드 대상 df 받아오기
        # 대상 파일들의 최근 종목코드 목록을 concat, 종목코드 의 중복제거
        for file_name in file_names:
            if df_targetCodesInfo.empty:
                df_targetCodesInfo = myWindow.readExcelToDataFrame(file_name, sheet_name)
            else:
                df_targetCodesInfo = pd.concat([df_targetCodesInfo, myWindow.readExcelToDataFrame(file_name, sheet_name)])
        df_targetCodesInfo.drop_duplicates(subset=['code'], keep='last') # code 중복 제거

        # 대상 종목코드 현재가 받아오기, 조건에 맞다면 매수 진행


        while True: # 9시 30분 까지 진행
            if datetime.datetime.now().hour == 9 and datetime.datetime.now().minute == 30:
                time.sleep(10)
                break

        while True: # 10 시 까지 또는 예수금이 모자르면 종료
            for idx, row in df_targetCodesInfo.iterrows():
                price_yesterday = row['종가']
                df_getStockInfo = myWindow.getStockInfo(row['code'], 2)
                now_price = df_getStockInfo.at[1, '종가']
                print("code = {0}, 현재가 = {1}, 전일종가 = {4}, 날짜 = {2}, 전일날짜 = {3}".format(row['code'], now_price,
                                                                                       df_getStockInfo.at[1, '날짜'],
                                                                                       sheet_name, price_yesterday))
                if True:
                    amount = buy_price // now_price
                    myWindow.buy_Stock(row['code'], amount, account_num)
                else:
                    pass
                time.sleep(0.2)
            if datetime.datetime.now().hour == 10 and datetime.datetime.now().minute == 0:
                break

        # myWindow.buy_Stock('005930', 1, '8038151811') # 매수
        # account_list = myWindow.get_account_info()
        # account_list = account_list.split(";")
        # account_list.remove("")
        # for i in account_list:
        #     print(i)
        #
        # myWindow.detail_account_mystock("8038151811", 0) # 계좌의 보유종목 들고오기

        # 코드리스트 정보 불러오기

        # while 9 < datetime.datetime.now().hour < 24:
        #     for i in use_code_list:
        #         time.sleep(0.5)
        #         df = myWindow.getStockInfo(i, 2)
        #         current_price = df.at[1, '종가']
        #         print(current_price)

                # 조건 일치시 매수진행


        # if datetime.datetime.now().hour < 19:
        #     os.system("shutdown -s -t 60")

        print("complete")
    except Exception as e:
        # if datetime.datetime.now().hour < 19:
        #     os.system("shutdown -s -t 60")
        print(e)
        print("Exception")

    # app.exec_()
